import os
import pandas as pd
import time
from app.services.step2_service import get_file_by_heuristic, get_col_strict
from app.services.automation_engine import PROJECT_ROOT, CACHE_DIR, get_cached_dataframe, get_audit_manager

def resolve_invoice_narration():
    """Step 4: Master Sync with Winning-Strategy Auto-Detection."""
    try:
        start_total = time.perf_counter()
        
        # 1. LOAD DATA
        z_path = os.path.join(CACHE_DIR, "Z_Recon_Step3.pkl")
        if not os.path.exists(z_path):
            return {"success": False, "error": "Step 3 Checkpoint not found."}
        
        z_df = pd.read_pickle(z_path)
        s_file = get_file_by_heuristic("SO Listing")
        so_df = get_cached_dataframe(s_file, engine='calamine')
        
        # 2. Identify Target Columns
        z_so_col = get_col_strict(z_df, "so no.")
        z_rev_col = get_col_strict(z_df, "revenue")
        if not z_so_col: return {"success": False, "error": "Missing SO Column."}

        # 3. IDENTIFY MASTER COLUMNS
        s_so1_col = get_col_strict(so_df, "sales order no", "sales order", "so no")
        s_inv_col = get_col_strict(so_df, "invoice no", "document number", "invoice")
        s_narr_col = get_col_strict(so_df, "narration", "text", "description")

        # --- AUTO-WINNER ENGINE (Selecting most accurate match rule) ---
        so_raw = so_df[s_so1_col].fillna("").astype(str).str.strip().str.upper()
        z_raw = z_df[z_so_col].fillna("").astype(str).str.strip().str.upper()
        
        # Strategy A: Clean (Our standard)
        so_a = so_raw.str.replace(r'\.0+$', '', regex=True).str.lstrip('0').replace("NAN", "")
        z_a = z_raw.str.replace(r'\.0+$', '', regex=True).str.lstrip('0').replace("NAN", "")
        map_a = dict(zip(so_a, so_df[s_inv_col].fillna("")))
        count_a = z_a.map(map_a).dropna().replace("", pd.NA).count()

        # Strategy B: Keep Decimal
        so_b = so_raw.str.lstrip('0').replace("NAN", "")
        z_b = z_raw.str.lstrip('0').replace("NAN", "")
        map_b = dict(zip(so_b, so_df[s_inv_col].fillna("")))
        count_b = z_b.map(map_b).dropna().replace("", pd.NA).count()

        # Strategy C: Keep Zeros
        so_c = so_raw.str.replace(r'\.0+$', '', regex=True).replace("NAN", "")
        z_c = z_raw.str.replace(r'\.0+$', '', regex=True).replace("NAN", "")
        map_c = dict(zip(so_c, so_df[s_inv_col].fillna("")))
        count_c = z_c.map(map_c).dropna().replace("", pd.NA).count()

        # Pick the winner
        best_count = max(count_a, count_b, count_c)
        if best_count == count_b:
            chosen_so, chosen_map, strategy_used = z_b, map_b, "Precision (Keep Decimals)"
        elif best_count == count_c:
            chosen_so, chosen_map, strategy_used = z_c, map_c, "Precision (Keep Zeros)"
        else:
            chosen_so, chosen_map, strategy_used = z_a, map_a, "Standard (Cleaned)"
        # -------------------------------------------------------------

        # 4. EXECUTE FINAL SYNC
        new_inv_col = "Invoice Number"
        new_narr_col = "Narration"
        if new_inv_col not in z_df.columns: z_df[new_inv_col] = ""
        if new_narr_col not in z_df.columns: z_df[new_narr_col] = ""

        narr_map = dict(zip(chosen_so, so_df[s_narr_col].fillna(""))) if s_narr_col else {}
        z_df[new_inv_col] = chosen_so.map(chosen_map).fillna("")
        if s_narr_col: z_df[new_narr_col] = chosen_so.map(narr_map).fillna("")

        success_indices = z_df[z_df[new_inv_col] != ""].index.tolist()
        total_targets = len(z_df[(z_df[z_so_col].notna()) & (z_df[z_rev_col] != 0)])

        z_df.to_pickle(os.path.join(CACHE_DIR, "Z_Recon_Step4.pkl"))

        # 5. Pipeline Audit Generator (RESTORED & IMPROVED)
        def save_master_sync_audit(df, root, updated_indices, inv_col, narr_col):
            t0 = time.perf_counter()
            from openpyxl import load_workbook
            from openpyxl.styles import PatternFill
            try:
                src_path = os.path.join(str(root), "Z_Recon_Standardized_Format.xlsx")
                if not os.path.exists(src_path): return
                
                wb = load_workbook(src_path)
                ws = wb.active
                headers = [str(cell.value).lower().strip() if cell.value else "" for cell in ws[1]]
                
                # Dynamic column update
                for col_name in [inv_col, narr_col]:
                    if col_name and col_name.lower().strip() not in headers:
                        ws.cell(row=1, column=ws.max_column+1).value = col_name
                        headers.append(col_name.lower().strip())
                
                inv_idx = headers.index(inv_col.lower().strip())
                narr_idx = headers.index(narr_col.lower().strip()) if narr_col.lower().strip() in headers else None
                green_fill = PatternFill(start_color='00FF00', end_color='00FF00', fill_type='solid')

                for idx in updated_indices:
                    row_num = df.index.get_loc(idx) + 2
                    ws.cell(row=row_num, column=inv_idx + 1).value = df.at[idx, inv_col]
                    ws.cell(row=row_num, column=inv_idx + 1).fill = green_fill
                    if narr_idx:
                        ws.cell(row=row_num, column=narr_idx + 1).value = df.at[idx, narr_col]
                        ws.cell(row=row_num, column=narr_idx + 1).fill = green_fill
                
                wb.save(src_path)
                print(f"🚀 [AuditManager] Step 4 Finish | {len(updated_indices)} rows in {int((time.perf_counter()-t0)*1000)}ms")
            except Exception: pass

        get_audit_manager().submit(save_master_sync_audit, z_df.copy(), PROJECT_ROOT, success_indices, new_inv_col, new_narr_col)

        duration = int((time.perf_counter() - start_total) * 1000)
        return {
            "success": True,
            "updates_applied": len(success_indices),
            "total_rows_to_check": total_targets,
            "execution_time_ms": duration,
            "process_steps": [
                {"label": "Winning Rule", "detail": f"Auto-Selected: {strategy_used}."},
                {"label": "Master Correlation", "detail": f"Matched {len(success_indices)} records to SO Listing."},
                {"label": "Final Audit", "detail": "Columns injected and cells highlighted Green."}
            ]
        }
    except Exception as e:
        return {"success": False, "error": str(e)}
