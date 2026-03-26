import os
import pandas as pd  # type: ignore
from typing import Dict, Any, Optional
from pathlib import Path
from config import BASE_DIR  # type: ignore
from app.services.automation_engine import (
    get_cached_dataframe, 
    normalize_sap_id, 
    get_col_from_df,
    INPUT_DIR,
    PROJECT_ROOT,
    CACHE_DIR
)

def get_system_files() -> list:
    """Returns the list of 5 required files tracked by the system (Strictly from Input Files)."""
    if not os.path.exists(INPUT_DIR):
        return []
    
    files = []
    for f in os.listdir(INPUT_DIR):
        if not f.startswith("~$") and "_Resolved" not in f and (f.lower().endswith('.xlsx') or f.lower().endswith('.xls')):
            try:
                card_type = "Z Recon" if "Z Recon" in f else "Revenue Dump" if "Revenue" in f else "Cost Dump" if "Cost" in f else "Invoice Listing" if "Invoice" in f else "SO Listing" if "SO" in f else None
                if card_type:
                    file_path = os.path.join(INPUT_DIR, f)
                    size = os.path.getsize(file_path)
                    files.append({
                        "name": f,
                        "size_kb": round(float(size / 1024.0), 1), # type: ignore
                        "type": card_type,
                        "status": "Ready",
                    })
            except Exception: pass
    return files

def get_file_by_heuristic(prefix: str):
    """Discovers a file across both Root and Input Files with fallback matching."""
    search_dirs = [INPUT_DIR, str(PROJECT_ROOT)]
    files_in_dir = []
    for d in search_dirs:
        if os.path.exists(d):
            files_in_dir.extend([os.path.join(d, f) for f in os.listdir(d) if not f.startswith("~$") and "_Resolved" not in f])
    
    if prefix == "Revenue Dump":
        file_path = next((f for f in files_in_dir if "revenue" in os.path.basename(f).lower()), None)
    elif prefix == "Cost":
        file_path = next((f for f in files_in_dir if "cost" in os.path.basename(f).lower()), None)
    else:
        file_path = next((f for f in files_in_dir if prefix.lower() in os.path.basename(f).lower()), None)
        
    if not file_path:
        raise Exception(f"Missing {prefix} source file.")
    return file_path

def get_col_strict(df, *keywords):
    """Enforce strict or partial column matching across variants."""
    for kw in keywords:
        for c in df.columns:
            if kw.lower() == str(c).lower().strip(): return c
    for kw in keywords:
        for c in df.columns:
            if kw.lower() in str(c).lower(): return c
    return None

def parse_zrecon() -> Dict[str, Any]:
    try:
        file_path = get_file_by_heuristic("Z Recon")
        z_df = get_cached_dataframe(file_path, engine='calamine')
        co_col = get_col_from_df(z_df, "company code")
        if co_col: z_df = z_df[z_df[co_col].notna()]
        rev_col = get_col_from_df(z_df, "revenue")
        cost_col = get_col_from_df(z_df, "cost")
        z_rev = pd.to_numeric(z_df[rev_col], errors='coerce').fillna(0).sum() if rev_col else 0.0
        z_cost = pd.to_numeric(z_df[cost_col], errors='coerce').fillna(0).sum() if cost_col else 0.0
        return {"success": True, "data": {"revenue": round(float(z_rev), 2), "cost": round(float(z_cost), 2)}}
    except Exception as e:
        return {"success": False, "error": str(e)}

def extract_programmatic_pivots(df: pd.DataFrame, source_type: str) -> tuple[float, int, bool, list]:
    col_str = {c: str(c).strip().lower() for c in df.columns}
    company_col = next((c for c, l in col_str.items() if "company" in l), None)
    material_col = next((c for c, l in col_str.items() if "material" in l), None)
    if source_type == "Revenue Dump":
        contract_col = next((c for c, l in col_str.items() if "reference key 3" in l), None) 
    else:
        contract_col = next((c for c, l in col_str.items() if "cc" == l or "text" in l), None)
    sum_col = get_col_from_df(df, "general ledger amount", "revenue", "cost", "value")
    if not sum_col: raise Exception(f"Math column not found for {source_type}.")
    df[sum_col] = pd.to_numeric(df[sum_col], errors='coerce').fillna(0)
    if company_col: df = df[df[company_col].notna()]
    totals = []; breakdowns = []
    def add_pivot(name, col):
        if not col: return
        grouped = df.groupby(col, dropna=False)[sum_col].sum()
        sum_v = float(grouped.sum())
        vals = [{"key": str(k), "amount": round(float(v), 2)} for k, v in grouped.items() if float(v) != 0.0]
        vals.sort(key=lambda x: abs(float(x["amount"])), reverse=True)
        breakdowns.append({"dimension": name, "total": round(sum_v, 2), "values": vals[:100]})
        totals.append(sum_v)
    add_pivot("Entity (Company Code)", company_col)
    add_pivot("Deputee (Material)", material_col)
    add_pivot("Contract Code (Ref Key 3 / Text)", contract_col)
    if totals:
        return totals[0], len(totals), all(abs(t - totals[0]) < 50 for t in totals), breakdowns
    return float(df[sum_col].sum()), 0, True, []

def parse_revenue() -> Dict[str, Any]:
    try:
        path = get_file_by_heuristic("Revenue Dump")
        df = get_cached_dataframe(path, sheet_name=1, engine='calamine')
        s, c, v, p = extract_programmatic_pivots(df, "Revenue Dump")
        return {"success": True, "data": {"revenue_sum": round(float(s), 2), "pivot_count": c, "pivots_consistent": v, "pivots": p}}
    except Exception as e: return {"success": False, "error": str(e)}

def parse_cost() -> Dict[str, Any]:
    try:
        path = get_file_by_heuristic("Cost")
        df = get_cached_dataframe(path, sheet_name=1, engine='calamine')
        s, c, v, p = extract_programmatic_pivots(df, "Cost Dump")
        return {"success": True, "data": {"cost_sum": round(float(s), 2), "pivot_count": c, "pivots_consistent": v, "pivots": p}}
    except Exception as e: return {"success": False, "error": str(e)}

def cross_invoice_integrity() -> Dict[str, Any]:
    """Step 2: Cross Invoice Integrity (VISUAL AUDIT EMPOWERED)"""
    try:
        z_path = get_file_by_heuristic("Z Recon")
        r_path = get_file_by_heuristic("Revenue Dump")
        i_path = get_file_by_heuristic("Invoice")
        
        z_df = get_cached_dataframe(z_path, engine='calamine')
        r_df = get_cached_dataframe(r_path, sheet_name=1, engine='calamine')
        if r_df.empty or len(r_df.columns) < 5: r_df = get_cached_dataframe(r_path, sheet_name=0, engine='calamine')
        i_df = get_cached_dataframe(i_path, engine='calamine')

        z_acc_col = get_col_strict(z_df, "accounting document")
        z_so_col = get_col_strict(z_df, "so no.", "so no", "so number", "sales order") or "SO Number"
        r_doc_col = get_col_strict(r_df, "document number", "doc. number", "accounting document")
        r_ref_col = get_col_strict(r_df, "reference key", "reference key 3", "invoice")
        i_inv_col = get_col_strict(i_df, "invoice no", "invoice", "billing document")
        i_so1_col = get_col_strict(i_df, "sales order no", "sales order", "sales order inv number")
        i_so2_col = get_col_strict(i_df, "so number2", "so number 2")

        if not all([z_acc_col, r_doc_col, r_ref_col, i_inv_col]):
            return {"success": False, "error": "Schema mismatch. Column AC or B/C missing."}

        # 1. identification
        if z_so_col not in z_df.columns: z_df[z_so_col] = None
        is_blank = z_df[z_so_col].isna() | (z_df[z_so_col].astype(str).str.strip().isin(['', 'nan']))
        
        acc_vec = z_df[z_acc_col].astype(str).str.strip().str.replace('.0','', regex=False, n=1).str.lstrip('0')
        mask_target = acc_vec.str.isdigit() & (~acc_vec.str.startswith('1'))
        target_indices = z_df[is_blank & mask_target].index

        # 2. Maps
        r_sub = r_df[[r_doc_col, r_ref_col]].dropna().copy()
        r_sub[r_doc_col] = r_sub[r_doc_col].astype(str).str.strip().str.replace('.0','',regex=False).str.lstrip('0')
        rev_map = r_sub.set_index(r_doc_col)[r_ref_col].to_dict()

        i_sub = i_df.dropna(subset=[i_inv_col]).copy()
        i_sub[i_inv_col] = i_sub[i_inv_col].astype(str).str.strip().str.replace('.0','',regex=False).str.lstrip('0')
        if i_so1_col and i_so2_col: i_sub['FINAL_SO'] = i_sub[i_so1_col].fillna(i_sub[i_so2_col])
        elif i_so1_col: i_sub['FINAL_SO'] = i_sub[i_so1_col]
        else: return {"success": False, "error": "No SO column in Invoice."}
        inv_map = i_sub.set_index(i_inv_col)['FINAL_SO'].to_dict()

        # 3. Resolve
        search_ids = acc_vec.loc[target_indices]
        resolved_refs = search_ids.map(rev_map).astype(str).str.strip().str.replace('.0','',regex=False).str.lstrip('0')
        final_sos = resolved_refs.map(inv_map)
        
        # Capture indices of success for highlighting
        success_indices = final_sos.dropna().index
        z_df.loc[target_indices, z_so_col] = final_sos.values
        
        # 4. Save with Highlight Audit Logic
        def save_highlighted_audit(df, root, updated_indices, so_col_name):
            import time
            t0 = time.perf_counter()
            try:
                import shutil
                from openpyxl import load_workbook
                from openpyxl.styles import PatternFill
                
                t_prep = time.perf_counter()
                # Source base file (Preserve original for audit)
                src_path = get_file_by_heuristic("Z Recon")
                out_path = os.path.join(str(root), "Z_Recon_Step2_Resolved.xlsx")
                
                # 4.1 Copy File
                shutil.copy2(src_path, out_path)
                t_copy = time.perf_counter()
                print(f"⏱️ [PERF] Audit: File Copied | Time: {int((t_copy - t_prep) * 1000)}ms")
                
                # 4.2 Load Workbook (THE WEIGHT)
                wb = load_workbook(out_path)
                ws = wb.active
                t_load = time.perf_counter()
                print(f"⏱️ [PERF] Audit: Workbook Loaded | Time: {int((t_load - t_copy) * 1000)}ms")
                
                # Identify column index
                headers = [str(cell.value).lower().strip() for cell in ws[1]]
                try:
                    col_idx = headers.index(so_col_name.lower().strip())
                except ValueError:
                    col_idx = df.columns.get_loc(so_col_name)

                yellow_fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')
                
                # 4.3 Update Loop
                t_loop_start = time.perf_counter()
                for idx in updated_indices:
                    row_pos = df.index.get_loc(idx)
                    val = df.at[idx, so_col_name]
                    target_cell = ws.cell(row=row_pos + 2, column=col_idx + 1)
                    target_cell.value = val
                    target_cell.fill = yellow_fill
                t_loop_end = time.perf_counter()
                print(f"⏱️ [PERF] Audit: Highlight Loop | Time: {int((t_loop_end - t_loop_start) * 1000)}ms ({len(updated_indices)} rows)")
                
                wb.save(out_path)
                t_final = time.perf_counter()
                print(f"🚀 [PERF] Audit: Total Save Complete | Time: {int((t_final - t0) * 1000)}ms")
            except Exception as e:
                print(f"⚠️ Optimized Audit Save Failed: {e}")

        # EXECUTE VIA PIPELINE QUEUE (Prevents File Locks across Steps)
        from app.services.automation_engine import get_audit_manager
        get_audit_manager().submit(
            save_highlighted_audit, 
            z_df.copy(), PROJECT_ROOT, success_indices, z_so_col
        )

        # Step 2 Checkpoint
        z_df.to_pickle(os.path.join(CACHE_DIR, "Z_Recon_Step2.pkl"))

        updates = int(len(success_indices))
        return {
            "success": True, 
            "updates_applied": updates,
            "unresolved_misses": int(len(target_indices) - updates),
            "total_rows_to_check": len(z_df),
            "process_steps": [
                {"label": "Direct Injection", "detail": f"Populated {updates} SOs into original Column: {z_so_col}."},
                {"label": "Smart Highlighter", "detail": f"Applied Yellow fill to {updates} newly resolved cells."},
                {"label": "Audit Generation", "detail": "Exporting 'Z_Recon_Step2_Resolved.xlsx' for manual review."}
            ]
        }
    except Exception as e: return {"success": False, "error": str(e)}

def validate_eager_all():
    """High-speed parallel orchestrator for the entire monthly pipeline."""
    from app.services.automation_engine import warmup_all_files
    import time
    
    start = time.time()
    # 1. Gather all Master Source file paths
    files_to_warm = []
    # Identify the 5 key master files
    for prefix in ["Z Recon", "Revenue Dump", "Cost dump", "SO Listing", "Invoice Listing"]:
        path = get_file_by_heuristic(prefix)
        if path: files_to_warm.append(path)
    
    # 2. Parallel Burst Warmup (Warms up binary cache for ALL steps)
    warmup_all_files(files_to_warm)
    
    # 3. Synchronous sub-results (Hitting cache now, so this is instant)
    z_res = parse_zrecon()
    r_res = parse_revenue()
    c_res = parse_cost()
    
    # 4. Aggregated result for the Dashboard
    return {
        "success": True,
        "execution_time_sec": round(time.time() - start, 2),
        "data": {
            "zrecon": z_res,
            "revenue": r_res,
            "cost": c_res
        },
        "warmed_up": [os.path.basename(f) for f in files_to_warm]
    }
