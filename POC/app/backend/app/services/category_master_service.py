import os
import pandas as pd
from openpyxl import load_workbook
from app.services.automation_engine import INPUT_DIR
from typing import List, Dict

MASTER_FILE_NAME = "Revenue file - Automation - Input Master.xlsx"
MASTER_FILE_PATH = os.path.join(INPUT_DIR, MASTER_FILE_NAME)
SHEET_NAME = "Input Master"

def get_category_mappings() -> List[Dict[str, str]]:
    """
    Reads the Category Master table from the Excel file.
    Table starts at row 2 (header 'CMIR type...'). Row 1 is the title.
    """
    if not os.path.exists(MASTER_FILE_PATH):
        return []

    # Read from row 2 (which is header)
    df = pd.read_excel(MASTER_FILE_PATH, sheet_name=SHEET_NAME, header=1, dtype=str)
    
    # We only care about the first two columns based on the image
    mappings = []
    # Columns are likely: Row 1 'CMIR type / Accounting Doc...' and Column 2 'Category'
    # pandas will read these as headers.
    col_a = df.columns[0]
    col_b = df.columns[1]
    
    for _, row in df.iterrows():
        cmir = str(row[col_a]).strip() if pd.notna(row[col_a]) else ""
        category = str(row[col_b]).strip() if pd.notna(row[col_b]) else ""
        if cmir:
            mappings.append({
                "type": cmir,
                "category": category
            })
            
    return mappings

def save_category_mappings(mappings: List[Dict[str, str]]):
    """
    Saves the mappings back to the Excel file, preserving the first two rows.
    """
    if not os.path.exists(MASTER_FILE_PATH):
        raise FileNotFoundError(f"Master file not found at {MASTER_FILE_PATH}")

    wb = load_workbook(MASTER_FILE_PATH)
    if SHEET_NAME not in wb.sheetnames:
        ws = wb.create_sheet(SHEET_NAME)
    else:
        ws = wb[SHEET_NAME]

    # Row 1: Title
    ws.cell(row=1, column=1).value = "5. Category Master"
    
    # Row 2: Headers
    ws.cell(row=2, column=1).value = "CMIR type / Accounting Doc number reference"
    ws.cell(row=2, column=2).value = "Category"

    # Row 3 onwards: Data
    # First, clear existing data from row 3 downwards (only first two columns)
    for row in ws.iter_rows(min_row=3, max_col=2):
        for cell in row:
            cell.value = None

    for i, m in enumerate(mappings):
        ws.cell(row=i+3, column=1).value = m.get("type", "")
        ws.cell(row=i+3, column=2).value = m.get("category", "")

    wb.save(MASTER_FILE_PATH)
    return True
