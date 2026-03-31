import os
import pandas as pd
import time
from app.services.automation_engine import PROJECT_ROOT, CACHE_DIR, get_audit_manager
from app.services.category_master_service import get_category_mappings

def execute_step6_category_mapping():
    """
    Step 6: Dynamic Category Mapping.
    
    Priority:
    1. Special Rule: If Accounting Document starts with '1', Category = 'Vendor manual'.
    2. Dynamic Rule: Match 'CMIR Type' from Step 3 against a user-managed master list.
    """
    try:
        start_total = time.perf_counter()
        
        # 1. LOAD DATA
        z_path = os.path.join(CACHE_DIR, "Z_Recon_Step5.pkl")
        if not os.path.exists(z_path):
            return {"success": False, "error": "Step 5 Checkpoint not found. Please run Step 5 first."}
        
        z_df = pd.read_pickle(z_path)
        
        # Identify Columns
        # Accounting Document (for Special Rule)
        acc_doc_col = None
        for col in z_df.columns:
            if "accounting document" in str(col).lower() or "document number" in str(col).lower():
                acc_doc_col = col
                break
        
        # CMIR Type (for Dynamic Mapping)
        cmir_type_col = None
        for col in z_df.columns:
            if "cmir type" in str(col).lower():
                cmir_type_col = col
                break
        
        # Target Category Column (Column AC usually, but we'll create/use 'Category')
        category_col = "Category"
        if category_col not in z_df.columns:
            z_df[category_col] = "" # Initialize if missing
            
        # 2. LOAD CATEGORY MASTER
        mapping_list = get_category_mappings()
        # Mapping dict: { 'type_string': 'category_value' }
        mapping_dict = { m["type"]: m["category"] for m in mapping_list }
        
        # Identifying the 'Special Rule' from the list if present
        starts_with_1_category = "Vendor manual" # Default fallback
        for k, v in mapping_dict.items():
            if "starting with 1 series" in k.lower():
                starts_with_1_category = v
                break

        # 3. APPLY CATEGORY LOGIC (WATERFALL)
        updates = 0
        
        def resolve_row_category(row):
            # Rule 1: Starts with 1 in Accounting Document
            doc_val = str(row.get(acc_doc_col, "")).strip()
            if doc_val.startswith("1"):
                return starts_with_1_category
            
            # Rule 2: Dynamic Match from CMIR Type
            cmir_val = str(row.get(cmir_type_col, "")).strip()
            if cmir_val and cmir_val in mapping_dict:
                return mapping_dict[cmir_val]
            
            return row.get(category_col, "")

        # Apply the mapping
        new_categories = z_df.apply(resolve_row_category, axis=1)
        
        # Track where we actually applied a new value compared to existing
        updated_indices = z_df[z_df[category_col] != new_categories].index
        z_df[category_col] = new_categories
        updates = len(updated_indices)
        
        # 4. SAVE CHECKPOINT & VISUAL AUDIT
        z_df.to_pickle(os.path.join(CACHE_DIR, "Z_Recon_Step6.pkl"))

        def save_step6_audit(df, root, modified_indices, col_name):
            from openpyxl import load_workbook
            from openpyxl.styles import PatternFill
            try:
                src_path = os.path.join(str(root), "Z_Recon_Standardized_Format.xlsx")
                if not os.path.exists(src_path): return
                
                wb = load_workbook(src_path)
                ws = wb.active
                
                # Find Column AC (We'll assume it's where 'Category' header either is or should be)
                headers = [str(cell.value).lower().strip() if cell.value else "" for cell in ws[1]]
                
                if col_name.lower() in headers:
                    col_num = headers.index(col_name.lower()) + 1
                else:
                    # Append it at the end if missing
                    col_num = ws.max_column + 1
                    ws.cell(row=1, column=col_num).value = col_name

                # Professional Light Cyan/Greenish highlight for Step 6
                step6_fill = PatternFill(start_color='E0FFFF', end_color='E0FFFF', fill_type='solid')

                for idx in modified_indices:
                    row_pos = df.index.get_loc(idx) + 2
                    val = df.at[idx, col_name]
                    cell = ws.cell(row=row_pos, column=col_num)
                    cell.value = val
                    cell.fill = step6_fill
                
                wb.save(src_path)
            except Exception as e:
                print(f"Audit Step 6 failed: {e}")

        # Submit to Background Worker
        get_audit_manager().submit(save_step6_audit, z_df.copy(), PROJECT_ROOT, updated_indices, category_col)

        duration = int((time.perf_counter() - start_total)*1000)
        
        return {
            "success": True,
            "updates_applied": updates,
            "execution_time_ms": duration,
            "message": f"Successfully mapped {updates} categories using Master Sync logic.",
            "process_steps": [
                {"label": "Standard Rule", "detail": f"Accounting documents starting with '1' assigned to '{starts_with_1_category}'."},
                {"label": "Master Rule Sync", "detail": f"Matched records against {len(mapping_list)} dynamic CMIR mappings."},
                {"label": "Audit Layout", "detail": f"Injected results into Category column with specialized highlighting."}
            ]
        }
    except Exception as e:
        import traceback
        print(traceback.format_exc())
        return {"success": False, "error": str(e)}
