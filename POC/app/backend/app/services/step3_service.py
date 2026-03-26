import os
import pandas as pd  # type: ignore
from typing import Dict, Any, Optional
from pathlib import Path
from config import BASE_DIR  # type: ignore
from app.services.automation_engine import (
    get_cached_dataframe, 
    normalize_sap_id, 
    get_col_from_df,
    INPUT_DIR,
    PROJECT_ROOT,
    CACHE_DIR
)
from app.services.step2_service import get_file_by_heuristic

def resolve_cmir_types() -> Dict[str, Any]:
    """Step 3: CMIR Type Resolution Flow (Waterfall Mapping)"""
    try:
        # 1. Load Step 2 Resovled State
        z_cache_path = os.path.join(CACHE_DIR, "Z_Recon_Step2.pkl")
        if not os.path.exists(z_cache_path):
            return {"success": False, "error": "Step 2 results not found. Please run Step 2 first."}
        
        z_df = pd.read_pickle(z_cache_path)
        
        # 2. Load SO Listing Source
        so_path = get_file_by_heuristic("SO Listing")
        so_df = get_cached_dataframe(so_path, engine='calamine')
        
        def get_col_strict(df, *keywords):
            for kw in keywords:
                for c in df.columns:
                    if kw.lower() == str(c).lower().strip(): return c
            for kw in keywords:
                for c in df.columns:
                    if kw.lower() in str(c).lower(): return c
            return None

        # 3. Target Identification in Z-Recon
        z_cmir_col = get_col_strict(z_df, "cmir type") or "CMIR Type"
        z_so_val_col = get_col_strict(z_df, "so no.", "so no", "so number")
        
        if not z_so_val_col:
            return {"success": False, "error": f"Cannot find SO Number column in Z-Recon to bridge matches."}

        # Identify rows where CMIR Typ is currently blank or 'nan'
        is_cmir_blank = z_df[z_cmir_col].isna() | (z_df[z_cmir_col].astype(str).str.strip().isin(['', 'nan']))
        target_indices = z_df[is_cmir_blank].index
        
        # 4. SO Listing Column Sniping
        s_so1_col = get_col_strict(so_df, "sales order no", "sales order")
        s_so2_col = get_col_strict(so_df, "so number2", "so number 2")
        s_trac_col = get_col_strict(so_df, "transaction type")
        
        if not s_trac_col:
            return {"success": False, "error": "Missing 'Transaction Type' (Col K) in SO Listing file."}
        if not s_so1_col:
            return {"success": False, "error": "Missing 'Sales Order No' (Col B) in SO Listing file."}

        # 5. Build Waterfall Cache (B -> K then Fallback SO2 -> K)
        # We build a single resolution map: normalize every key
        res_map = {}
        
        # Pass 1: Primary B (Sales Order No)
        s_sub1 = so_df.dropna(subset=[s_so1_col, s_trac_col]).copy()
        s_sub1["_KEY"] = s_sub1[s_so1_col].astype(str).str.strip().str.replace('.0','',regex=False).str.lstrip('0')
        res_map.update(s_sub1.set_index("_KEY")[s_trac_col].to_dict())
        
        # Pass 2: Fallback (SO Number2) - Does not overwrite existing map if they share a common key
        if s_so2_col:
            s_sub2 = so_df.dropna(subset=[s_so2_col, s_trac_col]).copy()
            s_sub2["_KEY"] = s_sub2[s_so2_col].astype(str).str.strip().str.replace('.0','',regex=False).str.lstrip('0')
            # Only update if the key isn't already there? Actually, user says "for rest cases", so we merge
            for k, v in s_sub2.set_index("_KEY")[s_trac_col].to_dict().items():
                if k not in res_map or pd.isna(res_map[k]):
                    res_map[k] = v

        # 6. High-Speed Merging
        search_ids = z_df.loc[target_indices, z_so_val_col].astype(str).str.strip().str.replace('.0','',regex=False).str.lstrip('0')
        final_trac_types = search_ids.map(res_map)
        
        success_indices = final_trac_types.dropna().index
        z_df.loc[success_indices, z_cmir_col] = final_trac_types.loc[success_indices].values
        
        # 7. Visual Audit: Highlighting (ORANGE)
        def save_orange_highlight_audit(df, root, updated_indices, cmir_col_name):
            try:
                out_path = os.path.join(str(root), "Z_Recon_Step2_Resolved.xlsx")
                col_idx = df.columns.get_loc(cmir_col_name)
                
                # We need to re-open the file we saved in Step 2 to preserve yellow highlights
                # Actually, writing OVER it is safer for consistency
                from openpyxl import load_workbook
                from openpyxl.styles import PatternFill
                
                # Check if file exists, if not, create standard
                if not os.path.exists(out_path):
                    df.to_excel(out_path, index=False)
                
                wb = load_workbook(out_path)
                ws = wb.active
                orange_fill = PatternFill(start_color='FFA500', end_color='FFA500', fill_type='solid')
                
                # 1. Map column name to Excel Column Letter/Index
                col_num = col_idx + 1
                
                # 2. Sync Values AND Highlights
                for idx in updated_indices:
                    row_pos = df.index.get_loc(idx)
                    val = df.at[idx, cmir_col_name]
                    
                    # Excel row = row_pos + 2 (1 for header, 1 for 1-indexed)
                    target_cell = ws.cell(row=row_pos + 2, column=col_num)
                    target_cell.value = val
                    target_cell.fill = orange_fill
                
                wb.save(out_path)
                print(f"✅ Visual Step 3 Audit Sync Complete: {out_path}")
            except Exception as e:
                print(f"⚠️ Failed Step 3 Visual Audit: {e}")

        # EXECUTE VIA PIPELINE QUEUE (Serialized to avoid file locks)
        from app.services.automation_engine import get_audit_manager
        get_audit_manager().submit(
            save_orange_highlight_audit, 
            z_df.copy(), PROJECT_ROOT, success_indices, z_cmir_col
        )
        
        # SAVE CHECKPOINT FOR STEP 4
        z_df.to_pickle(os.path.join(CACHE_DIR, "Z_Recon_Step3.pkl"))
        
        updates = int(len(success_indices))
        return {
            "success": True, 
            "updates_applied": updates,
            "unresolved_misses": int(len(target_indices) - updates),
            "total_rows_to_check": len(target_indices),
            "process_steps": [
                {"label": "Gap Analysis", "detail": f"Found {len(target_indices)} blank fields in Column E (CMIR Type)."},
                {"label": "Master Source Sync", "detail": f"Warmed up {len(so_df)} entries from SO Listing workbook."},
                {"label": "Waterfall Bridging", "detail": f"Built a resolution map with {len(res_map)} Transaction types."},
                {"label": "Transaction Type Injection", "detail": f"Successfully mapped {updates} records across files."},
                {"label": "Self-Audit Logic", "detail": "Applying Orange cell-fills to Column E in audit Excel."}
            ]
        }
    except Exception as e:
        return {"success": False, "error": str(e)}
