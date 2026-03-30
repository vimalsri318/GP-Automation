import os
import pandas as pd
import time
from app.services.step2_service import get_file_by_heuristic, get_col_strict
from app.services.automation_engine import PROJECT_ROOT, CACHE_DIR, get_cached_dataframe, get_audit_manager

def resolve_secondary_narration():
    """
    Step 5: Secondary Narration Recovery Flow (Strict Alignment).
    
    Logic:
    Phase 1: (SO exists, Narration is blank)
       - Match with SO Listing (Sales Order No or SO Number2).
       - Take Narration column from SO Listing.
    Phase 2: (Both SO and Narration are unavailable)
       - Match Accounting Document with Document number in Cost Dump (Sheet1).
       - Take 'Text' column as Narration.
    """
    try:
        start_total = time.perf_counter()
        
        # 1. LOAD DATA
        z_path = os.path.join(CACHE_DIR, "Z_Recon_Step4.pkl")
        if not os.path.exists(z_path):
            return {"success": False, "error": "Step 4 Checkpoint not found."}
        
        z_df = pd.read_pickle(z_path)
        
        # Identify Target Columns in Z-Recon
        z_so_col = get_col_strict(z_df, "so no.", "so no", "so number", "sales order")
        z_acc_col = get_col_strict(z_df, "accounting_document", "accounting document", "document number")
        z_narr_col = "Narration" # Created in Step 4
        
        if z_narr_col not in z_df.columns:
            z_df[z_narr_col] = ""
            
        # Identify rows with empty/null narration
        is_narr_blank = z_df[z_narr_col].isna() | (z_df[z_narr_col].astype(str).str.strip().isin(['', 'nan']))
        target_indices = z_df[is_narr_blank].index
        
        if len(target_indices) == 0:
            return {
                "success": True, 
                "updates_applied": 0, 
                "message": "All narrations already populated.",
                "process_steps": [{"label": "Status", "detail": "No blanks found in Narration column."}]
            }

        all_new_indices = []
        
        # 2. PHASE 1: SO List (Condition: SO available, Narration blank)
        # ---------------------------------------------------------------------
        z_with_so_indices = z_df.loc[target_indices][z_df.loc[target_indices][z_so_col].notna() & ~(z_df.loc[target_indices][z_so_col].astype(str).str.strip().isin(['', 'nan']))].index
        updates_a = 0
        
        if len(z_with_so_indices) > 0:
            s_file = get_file_by_heuristic("SO Listing")
            so_df = get_cached_dataframe(s_file, engine='calamine')
            
            s_so1_col = get_col_strict(so_df, "sales order no", "sales order", "so no")
            s_so2_col = get_col_strict(so_df, "so number2", "so number 2")
            s_narr_col = get_col_strict(so_df, "narration", "text", "description")
            
            if s_narr_col:
                so_map = {}
                # Pass 1: Primary SO
                s_sub1 = so_df.dropna(subset=[s_so1_col, s_narr_col]).copy()
                s_sub1["_KEY"] = s_sub1[s_so1_col].astype(str).str.strip().str.replace(r'\.0+$', '', regex=True).str.lstrip('0')
                so_map.update(s_sub1.set_index("_KEY")[s_narr_col].to_dict())
                
                # Pass 2: Fallback SO2
                if s_so2_col:
                    s_sub2 = so_df.dropna(subset=[s_so2_col, s_narr_col]).copy()
                    s_sub2["_KEY"] = s_sub2[s_so2_col].astype(str).str.strip().str.replace(r'\.0+$', '', regex=True).str.lstrip('0')
                    for k, v in s_sub2.set_index("_KEY")[s_narr_col].to_dict().items():
                        if k not in so_map or pd.isna(so_map[k]):
                            so_map[k] = v
                
                search_ids = z_df.loc[z_with_so_indices, z_so_col].astype(str).str.strip().str.replace(r'\.0+$', '', regex=True).str.lstrip('0')
                resolved_narrs = search_ids.map(so_map)
                
                success_indices_a = resolved_narrs.dropna().index
                z_df.loc[success_indices_a, z_narr_col] = resolved_narrs.loc[success_indices_a].values
                updates_a = len(success_indices_a)
                all_new_indices.extend(success_indices_a.tolist())

        # 3. PHASE 2: Cost Dump (Condition: BOTH SO and Narration are unavailable)
        # ---------------------------------------------------------------------
        is_so_blank = z_df[z_so_col].isna() | (z_df[z_so_col].astype(str).str.strip().isin(['', 'nan']))
        is_narr_blank_now = z_df[z_narr_col].isna() | (z_df[z_narr_col].astype(str).str.strip().isin(['', 'nan']))
        remaining_indices = z_df[is_so_blank & is_narr_blank_now].index
        updates_b = 0
        
        if len(remaining_indices) > 0:
            c_file = get_file_by_heuristic("Cost")
            # Logic strictly uses 'Sheet1' per user request
            cost_df = get_cached_dataframe(c_file, sheet_name='Sheet1', engine='calamine')
            
            c_doc_col = get_col_strict(cost_df, "document number", "doc. number", "accounting document")
            c_text_col = get_col_strict(cost_df, "text", "narration", "description")
            
            if c_doc_col and c_text_col:
                cost_map = cost_df.dropna(subset=[c_doc_col, c_text_col]).copy()
                cost_map[c_doc_col] = cost_map[c_doc_col].astype(str).str.strip().str.replace(r'\.0+$', '', regex=True).str.lstrip('0')
                c_resolver = cost_map.set_index(c_doc_col)[c_text_col].to_dict()
                
                search_docs = z_df.loc[remaining_indices, z_acc_col].astype(str).str.strip().str.replace(r'\.0+$', '', regex=True).str.lstrip('0')
                resolved_costs = search_docs.map(c_resolver)
                
                success_indices_b = resolved_costs.dropna().index
                z_df.loc[success_indices_b, z_narr_col] = resolved_costs.loc[success_indices_b].values
                updates_b = len(success_indices_b)
                all_new_indices.extend(success_indices_b.tolist())

        # 4. SAVE CHECKPOINT & LOG RECOVERED SAMPLES
        z_df.to_pickle(os.path.join(CACHE_DIR, "Z_Recon_Step5.pkl"))
        
        # Build Sample Preview (First 5 recovered narrations)
        samples = []
        if len(all_new_indices) > 0:
            sample_df = z_df.loc[all_new_indices[:5]]
            for idx, row in sample_df.iterrows():
                samples.append({
                    "id": str(row.get(z_acc_col, idx)),
                    "narration": str(row[z_narr_col])
                })
        
        # 5. VISUAL AUDIT (REFUNED PURPLE HIGHLIGHTS)
        def save_step5_audit_strict(df, root, newly_updated_indices, narr_col):
            from openpyxl import load_workbook
            from openpyxl.styles import PatternFill
            try:
                src_path = os.path.join(str(root), "Z_Recon_Standardized_Format.xlsx")
                if not os.path.exists(src_path): return
                
                wb = load_workbook(src_path)
                ws = wb.active
                headers = [str(cell.value).lower().strip() if cell.value else "" for cell in ws[1]]
                
                if narr_col.lower().strip() not in headers:
                    ws.cell(row=1, column=ws.max_column+1).value = narr_col
                    headers.append(narr_col.lower().strip())
                
                narr_idx = headers.index(narr_col.lower().strip())
                step5_fill = PatternFill(start_color='DDA0DD', end_color='DDA0DD', fill_type='solid')

                for idx in newly_updated_indices:
                    row_num = df.index.get_loc(idx) + 2
                    ws.cell(row=row_num, column=narr_idx + 1).value = df.at[idx, narr_col]
                    ws.cell(row=row_num, column=narr_idx + 1).fill = step5_fill
                
                wb.save(src_path)
            except Exception: pass

        get_audit_manager().submit(save_step5_audit_strict, z_df.copy(), PROJECT_ROOT, all_new_indices, z_narr_col)

        duration = int((time.perf_counter() - start_total) * 1000)
        return {
            "success": True,
            "updates_applied": updates_a + updates_b,
            "updates_via_so": updates_a,
            "updates_via_cost": updates_b,
            "execution_time_ms": duration,
            "samples": samples,
            "process_steps": [
                {"label": "SO Recovery (Phase 1)", "detail": f"Injected {updates_a} narrations using SO Mapping."},
                {"label": "Cost Fallback (Phase 2)", "detail": f"Injected {updates_b} narrations using Accounting Doc (Sheet1)."},
                {"label": "UI Visibility", "detail": f"Recovered {len(all_new_indices)} items. Samples available in dashboard."}
            ]
        }
    except Exception as e:
        import traceback
        print(traceback.format_exc())
        return {"success": False, "error": str(e)}
